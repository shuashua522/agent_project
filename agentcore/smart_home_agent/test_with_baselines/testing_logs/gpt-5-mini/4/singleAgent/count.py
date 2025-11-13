import re
import os
import openpyxl
from openpyxl.styles import Font


def count_target_fields(log_file_path):
    # 定义目标字段及对应的正则表达式（严格匹配格式要求）
    field_patterns = {
        'entity_id': r'"entity_id":\s*"[a-zA-Z0-9_.]+',  # 后跟字符串格式（字母/数字/下划线/点）
        'state': r'"state":\s*"[a-zA-Z0-9_.]+',  # 后跟字符串格式
        'volume_level': r'"volume_level":\s*\d+(\.\d+)?',  # 冒号+数字（整数或小数）
        'brightness': r'"brightness":\s*\d+(\.\d+)?',  # 冒号+数字
        'color_temp_kelvin': r'"color_temp_kelvin":\s*\d+(\.\d+)?',  # 冒号+数字

        # 新增字段（均匹配：字段名: "字符串内容" 格式）
        'last_changed': r'"last_changed":\s*"[^"]+"',  # 冒号+带引号的字符串（支持日期时间等格式）
        'last_reported': r'"last_reported":\s*"[^"]+"',
        'last_updated': r'"last_updated":\s*"[^"]+"',
        'id': r'"id":\s*"[^"]+"',  # 匹配id: "xxx" 格式（字符串内容可为任意非引号字符）
        'parent_id': r'"parent_id":\s*"[^"]+"',
        'user_id': r'"user_id":\s*"[^"]+"'
    }

    # 初始化计数字典
    counts = {field: 0 for field in field_patterns.keys()}

    # 标记是否进入Tool Message部分
    in_tool_message = False

    try:
        # 读取日志文件
        with open(log_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                # 检测Tool Message开始标记
                if '================================= Tool Message =================================' in line:
                    in_tool_message = True
                    continue
                # 检测下一个消息类型（结束Tool Message部分）
                elif '================================= Ai Message ==================================' in line or \
                        '================================ Human Message =================================' in line:
                    in_tool_message = False
                    continue

                # 只处理Tool Message部分的内容
                if in_tool_message:
                    # 匹配每个目标字段并计数
                    for field, pattern in field_patterns.items():
                        matches = re.findall(pattern, line)
                        counts[field] += len(matches)

    except FileNotFoundError:
        print(f"错误：未找到文件 {log_file_path}")
        return None
    except Exception as e:
        print(f"处理文件 {log_file_path} 时出错：{str(e)}")
        return None

    return counts


def save_to_excel(results, fields):
    """将统计结果保存到Excel文件"""
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "隐私计数结果"

    # 设置标题行
    header = ["文件序号"] + fields
    for col, title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = title
        cell.font = Font(bold=True)  # 标题加粗

    # 按文件序号排序并写入数据
    sorted_files = sorted(results.keys(), key=lambda x: int(x.split("_")[0]))
    for row, file_name in enumerate(sorted_files, 2):
        # 写入文件序号
        file_num = file_name.split("_")[0]
        ws.cell(row=row, column=1).value = file_num

        # 写入各字段计数
        counts = results[file_name]
        for col, field in enumerate(fields, 2):
            ws.cell(row=row, column=col).value = counts[field]

    # 调整列宽
    for col in range(1, len(header) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # 保存文件
    try:
        wb.save("隐私计数结果.xlsx")
        print("结果已成功保存到 '隐私计数结果.xlsx'")
    except Exception as e:
        print(f"保存Excel文件时出错：{str(e)}")


def main():
    # 定义字段顺序（与要求的标题顺序一致）
    fields_order = [
        'entity_id', 'state', 'volume_level', 'brightness',
        'color_temp_kelvin', 'last_changed', 'last_reported',
        'last_updated', 'id', 'parent_id', 'user_id'
    ]

    # 存储所有文件的统计结果
    all_results = {}
    # 遍历1到50的序号
    for num in range(1, 51):
        # 构建可能的文件名模式
        file_pattern = f"{num}_*.log"
        # 在当前目录查找匹配的文件
        matching_files = [f for f in os.listdir('.') if os.path.isfile(f) and re.match(f"^{num}_.*\.log$", f)]

        if matching_files:
            # 假设每个序号对应唯一的测试用例文件
            log_file = matching_files[0]
            print(f"正在处理文件: {log_file}")
            results = count_target_fields(log_file)
            if results:
                all_results[log_file] = results
        else:
            print(f"未找到序号为 {num} 的日志文件")

    # 打印并保存统计结果
    if all_results:
        # 打印结果
        print("\n所有文件目标字段统计结果汇总：")
        for file_name, counts in all_results.items():
            print(f"\n文件: {file_name}")
            for field, count in counts.items():
                print(f"{field}: {count} 次")

        # 保存到Excel
        save_to_excel(all_results, fields_order)
    else:
        print("未统计到任何有效文件的结果")


if __name__ == "__main__":
    main()