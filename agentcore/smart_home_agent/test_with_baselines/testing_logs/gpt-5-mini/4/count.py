import re
import os
import openpyxl
from openpyxl.styles import Font


def count_target_fields(log_file_path):
    # 定义目标字段及对应的正则表达式（严格匹配格式要求）
    field_patterns = {
        # 原有数值型字段不变
        'volume_level': r'"volume_level":\s*\d+(\.\d+)?',  # 冒号+数字（整数或小数）
        'brightness': r'"brightness":\s*\d+(\.\d+)?',  # 冒号+数字
        'color_temp_kelvin': r'"color_temp_kelvin":\s*\d+(\.\d+)?',  # 冒号+数字

        # 字符串型字段：排除被@包裹的情况（简化否定逻辑）
        'entity_id': r'"entity_id":\s*"[^@"]+"',  # 字符串中无@才匹配
        'state': r'"state":\s*"[^@"]+"',
        'last_changed': r'"last_changed":\s*"[^@"]+"',
        'last_reported': r'"last_reported":\s*"[^@"]+"',
        'last_updated': r'"last_updated":\s*"[^@"]+"',
        'id': r'"id":\s*"[^@"]+"',
        'user_id': r'"user_id":\s*"[^@"]+"',
        # parent_id：仅匹配带引号且不含@的字符串，忽略null
        'parent_id': r'"parent_id":\s*"[^@"]+"'  # 只统计"xxx"格式，null不计数
    }
    """
    field_patterns = {
        'entity_id': r'"entity_id":\s*"[a-zA-Z0-9_.]+',  # 后跟字符串格式（字母/数字/下划线/点）
        'state': r'"state":\s*"[a-zA-Z0-9_.]+',  # 后跟字符串格式
        'volume_level': r'"volume_level":\s*\d+(\.\d+)?',  # 冒号+数字（整数或小数）
        'brightness': r'"brightness":\s*\d+(\.\d+)?',  # 冒号+数字
        'color_temp_kelvin': r'"color_temp_kelvin":\s*\d+(\.\d+)?',  # 冒号+数字

        # 新增字段（均匹配：字段名: "字符串内容" 格式）
        'last_changed': r'"last_changed":\s*"[^"]+"',  # 冒号+带引号的字符串（支持日期时间等格式）
        'last_reported': r'"last_reported":\s*"[^"]+"',
        'last_updated': r'"last_updated":\s*"[^"]+"',
        'id': r'"id":\s*"[^"]+"',  # 匹配id: "xxx" 格式（字符串内容可为任意非引号字符）
        'parent_id': r'"parent_id":\s*"[^"]+"',
        'user_id': r'"user_id":\s*"[^"]+"'
    }
    """

    # 初始化计数字典
    counts = {field: 0 for field in field_patterns.keys()}

    # 标记是否进入Tool Message部分
    in_tool_message = False

    try:
        # 读取日志文件
        with open(log_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                # 检测Tool Message开始标记
                if '================================= Tool Message =================================' in line:
                    in_tool_message = True
                    continue
                # 检测下一个消息类型（结束Tool Message部分）
                elif '================================= Ai Message ==================================' in line or \
                        '================================ Human Message =================================' in line:
                    in_tool_message = False
                    continue

                # 只处理Tool Message部分的内容
                if in_tool_message:
                    # 匹配每个目标字段并计数
                    for field, pattern in field_patterns.items():
                        # matches = re.findall(pattern, line)
                        # counts[field] += len(matches)
                        # 关键修复：用re.finditer统计完整匹配次数，避免分组干扰
                        matches = re.finditer(pattern, line)
                        counts[field] += sum(1 for _ in matches)  # 计数完整匹配的次数


    except FileNotFoundError:
        print(f"错误：未找到文件 {log_file_path}")
        return None
    except Exception as e:
        print(f"处理文件 {log_file_path} 时出错：{str(e)}")
        return None

    return counts


def save_to_excel(agent_results, fields):
    """将各agent的统计结果保存到Excel文件的不同工作表中"""
    # 创建工作簿
    wb = openpyxl.Workbook()
    # 删除默认创建的Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    for agent, results in agent_results.items():
        # 为每个agent创建工作表
        ws = wb.create_sheet(title=agent)

        # 设置标题行
        header = ["文件序号"] + fields
        for col, title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = title
            cell.font = Font(bold=True)  # 标题加粗

        # 按文件序号排序并写入数据
        sorted_files = sorted(results.keys(), key=lambda x: int(x.split("_")[0]))
        for row, file_name in enumerate(sorted_files, 2):
            # 写入文件序号
            file_num = file_name.split("_")[0]
            ws.cell(row=row, column=1).value = file_num

            # 写入各字段计数
            counts = results[file_name]
            for col, field in enumerate(fields, 2):
                ws.cell(row=row, column=col).value = counts[field]

        # 调整列宽
        for col in range(1, len(header) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # 保存文件
    try:
        wb.save("隐私计数结果.xlsx")
        print("结果已成功保存到 '隐私计数结果.xlsx'")
    except Exception as e:
        print(f"保存Excel文件时出错：{str(e)}")


def main():
    # 定义需要统计的agent目录
    agents = ['singleAgent', 'sashaAgent', 'sageAgent', 'privacyAgent']
    # 定义字段顺序
    fields_order = [
        'entity_id', 'state', 'volume_level', 'brightness',
        'color_temp_kelvin', 'last_changed', 'last_reported',
        'last_updated', 'id', 'parent_id', 'user_id'
    ]

    # 存储所有agent的统计结果，结构: {agent: {filename: counts}}
    all_agent_results = {agent: {} for agent in agents}

    for agent in agents:
        # 检查目录是否存在
        if not os.path.isdir(agent):
            print(f"警告：目录 {agent} 不存在，将跳过该agent的统计")
            continue

        print(f"\n开始处理 {agent} 目录...")
        # 遍历1到50的序号
        for num in range(1, 51):
            # 构建文件名匹配模式
            file_pattern = f"{num}_*.log"
            # 获取目录下所有文件
            dir_files = [f for f in os.listdir(agent) if os.path.isfile(os.path.join(agent, f))]
            # 匹配符合条件的日志文件
            matching_files = [f for f in dir_files if re.match(f"^{num}_.*\.log$", f)]

            if matching_files:
                log_file = matching_files[0]
                log_file_path = os.path.join(agent, log_file)
                print(f"正在处理文件: {log_file_path}")
                results = count_target_fields(log_file_path)
                if results:
                    all_agent_results[agent][log_file] = results
            else:
                print(f"{agent} 目录下未找到序号为 {num} 的日志文件")

    # 保存到Excel
    save_to_excel(all_agent_results, fields_order)


if __name__ == "__main__":
    main()